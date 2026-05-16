"""Excel report generation using openpyxl."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.reports.report_schema import (
    DEFECT_CANDIDATE_FIELDS,
    IMAGE_RESULT_FIELDS,
    MISCLASSIFIED_FIELDS,
)


class ExcelReport:
    """Generate multi-sheet Excel reports with formatted data."""

    def __init__(self, output_path: str | Path) -> None:
        self.output_path = Path(output_path)
        self.wb = openpyxl.Workbook()
        self._header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        self._header_font = Font(bold=True, size=11, color="FFFFFF")
        self._bold_font = Font(bold=True)
        self._thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self._header_alignment = Alignment(horizontal="center", vertical="center")

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _write_header(self, ws, headers: list[str]) -> None:
        """Write styled header row."""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._header_alignment
            cell.border = self._thin_border

    def _write_data_row(self, ws, row_idx: int, values: list[Any]) -> None:
        """Write a data row with thin borders."""
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = self._thin_border
            cell.alignment = Alignment(vertical="center")

    def _auto_width(self, ws, min_width: int = 8, max_width: int = 55) -> None:
        """Auto-size columns based on content."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            adjusted = max(min_width, min(max_len + 2, max_width))
            ws.column_dimensions[col_letter].width = adjusted

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def add_summary_sheet(self, metrics: dict[str, Any]) -> None:
        """Add a summary sheet with key metrics.

        Args:
            metrics: Dict with keys matching SUMMARY_FIELDS and additional metadata.
        """
        ws = self.wb.active
        ws.title = "Summary"

        self._write_header(ws, ["Metric", "Value"])

        rows = [
            ("Total Images", metrics.get("total_images", 0)),
            ("OK False Positive Rate", f"{metrics.get('ok_fpr', 0):.4f}"),
            ("NG Miss Rate", f"{metrics.get('ng_miss_rate', 0):.4f}"),
            ("Acceptable Micro Defect FP Rate", f"{metrics.get('acceptable_micro_fpr', 0):.4f}"),
            ("Unknown Defect Recall", f"{metrics.get('unknown_recall', 0):.4f}"),
            ("Borderline Detection Rate", f"{metrics.get('borderline_detection_rate', 0):.4f}"),
            ("Average Inference Time (ms)", f"{metrics.get('avg_inference_time_ms', 0):.1f}"),
            ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for r, (metric, value) in enumerate(rows, 2):
            cell_metric = ws.cell(row=r, column=1, value=metric)
            cell_metric.font = self._bold_font
            cell_metric.border = self._thin_border

            cell_value = ws.cell(row=r, column=2, value=value)
            cell_value.border = self._thin_border

        self._auto_width(ws)

    def add_image_results_sheet(self, records: list[Any]) -> None:
        """Add a sheet with per-image fusion results.

        Args:
            records: List of ImageRecord objects.
        """
        ws = self.wb.create_sheet("Image Results")
        self._write_header(ws, IMAGE_RESULT_FIELDS)

        for i, rec in enumerate(records, 2):
            fd = rec.fusion_decision

            yolo_str = (
                str(len(rec.yolo_result.predictions))
                if rec.yolo_result and rec.yolo_result.predictions
                else "0"
            )
            patchcore_score = (
                round(rec.patchcore_result.anomaly.image_score, 4)
                if rec.patchcore_result and rec.patchcore_result.anomaly
                else "N/A"
            )
            efficientad_score = (
                round(rec.efficientad_result.anomaly.image_score, 4)
                if rec.efficientad_result and rec.efficientad_result.anomaly
                else "N/A"
            )
            fastflow_score = (
                round(rec.fastflow_result.anomaly.image_score, 4)
                if rec.fastflow_result and rec.fastflow_result.anomaly
                else "N/A"
            )
            opencv_str = (
                str(len(rec.opencv_result.predictions))
                if rec.opencv_result and rec.opencv_result.predictions
                else "0"
            )

            self._write_data_row(
                ws,
                i,
                [
                    rec.image_path,
                    rec.true_label,
                    yolo_str,
                    patchcore_score,
                    efficientad_score,
                    fastflow_score,
                    opencv_str,
                    fd.strategy.value if fd else "N/A",
                    fd.final_decision.value if fd else "N/A",
                    fd.reason if fd else "N/A",
                    "Misclassified" if rec.is_misclassified else "Correct",
                    round(fd.runtime_ms, 1) if fd else 0,
                ],
            )

        self._auto_width(ws)

    def add_defect_candidates_sheet(self, all_candidates: list[Any]) -> None:
        """Add a sheet listing all individual defect candidates.

        Args:
            all_candidates: List of DefectCandidate objects.
        """
        ws = self.wb.create_sheet("Defect Candidates")
        self._write_header(ws, DEFECT_CANDIDATE_FIELDS)

        for i, c in enumerate(all_candidates, 2):
            bbox_str = f"[{c.bbox_xyxy[0]:.0f}, {c.bbox_xyxy[1]:.0f}, {c.bbox_xyxy[2]:.0f}, {c.bbox_xyxy[3]:.0f}]"
            self._write_data_row(
                ws,
                i,
                [
                    c.image_path,
                    c.candidate_id,
                    c.source_model.value,
                    c.class_name,
                    round(c.confidence, 4),
                    bbox_str,
                    round(c.area_px, 1),
                    round(c.length_px, 1),
                    round(c.width_px, 1),
                    round(c.area_mm2, 4) if c.area_mm2 is not None else "N/A",
                    round(c.length_mm, 4) if c.length_mm is not None else "N/A",
                    round(c.width_mm, 4) if c.width_mm is not None else "N/A",
                    round(c.aspect_ratio, 2),
                    round(c.max_anomaly_score, 4),
                    "N/A",
                    "N/A",
                ],
            )

        self._auto_width(ws)

    def add_misclassified_sheet(self, misclassified: list[Any]) -> None:
        """Add a sheet with misclassified samples for review.

        Args:
            misclassified: List of ImageRecord objects flagged as misclassified.
        """
        ws = self.wb.create_sheet("Misclassified Samples")
        self._write_header(ws, MISCLASSIFIED_FIELDS)

        for i, rec in enumerate(misclassified, 2):
            fd = rec.fusion_decision
            self._write_data_row(
                ws,
                i,
                [
                    rec.image_path,
                    rec.true_label,
                    fd.final_decision.value if fd else "N/A",
                    rec.error_type,
                    fd.reason if fd else "N/A",
                ],
            )

        self._auto_width(ws)

    def add_strategy_comparison_sheet(self, comparison: list[dict[str, Any]]) -> None:
        """Add a sheet comparing multiple fusion strategies.

        Args:
            comparison: List of dicts from compute_strategy_comparison().
        """
        ws = self.wb.create_sheet("Strategy Comparison")

        if not comparison:
            ws.cell(row=1, column=1, value="No comparison data available").font = self._bold_font
            return

        headers = list(comparison[0].keys())
        self._write_header(ws, headers)

        for i, row_data in enumerate(comparison, 2):
            self._write_data_row(ws, i, [row_data.get(k, "") for k in headers])

        self._auto_width(ws)

    def save(self) -> str:
        """Write the workbook to disk and return the output path."""
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(self.output_path))
        return str(self.output_path)
